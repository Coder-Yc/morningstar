from openpyxl import load_workbook
import openpyxl
import os
import re
import json
import os
import time
from classes.GetMorningStarDataClass_vs1 import GetMorningStarData

countries = ['Shanghai_Shenzhen']
# check_Path = "./RawData/check.xlsx"

if __name__ == "__main__":
    rb = load_workbook("check.xlsx")
    sheets_names = rb.get_sheet_names()
    if "TSX" not in sheets_names:
        rb.create_sheet('TSX')
    if "Snp500_Ru1000" not in sheets_names:
        rb.create_sheet('Snp500_Ru1000')
    if "Shanghai_Shenzhen" not in sheets_names:
        rb.create_sheet("Shanghai_Shenzhen")
    # if "Sheet1" in sheets_names:
    #     rb.remove_sheet("Sheet1")
    rb.save("check.xlsx")
    for country in countries:
        wb = load_workbook("check.xlsx")
        wb['{}'.format(country)]['A1'] = 'symbol'
        wb['{}'.format(country)]['B1'] = 'download'
        wb['{}'.format(country)]['C1'] = 'Tab'
        wb['{}'.format(country)]['D1'] = 'download sucessfully'
        wb.save("check.xlsx")
        path = "./RawData/{}".format(country)
        index = 2

        print("=======开始检查========")
        for file in os.listdir(path):
            print(file)
            find_ = re.compile(r"Download_(.*?).xlsx")
            symbol = find_.findall(file)
            symbol_path = "./RawData/{}/{}".format(country,file)
            file = openpyxl.load_workbook(symbol_path)
            file_tag = len(file.sheetnames)
            file_size = os.path.getsize(symbol_path)
            if file_size>= 3275:
                download = "yes"
            else:
                download = "no"
            if download == "yes" and file_tag == 8:
                download_sucessfully = "yes"
            else:
                download_sucessfully = "no"

            readbook = load_workbook("check.xlsx")
            sheet = readbook['{}'.format(country)]
            sheet.cell(column=1,row=index,value=symbol[0])
            sheet.cell(column=2,row=index,value=download)
            sheet.cell(column=3,row=index,value=file_tag)
            sheet.cell(column=4,row=index,value=download_sucessfully)
            index+=1
            readbook.save("check.xlsx")
        print("========检查结束========")
        readbook = load_workbook("check.xlsx")
        sheet = readbook['{}'.format(country)]
        col_1 = sheet["D"]
        col_2 = sheet["A"]
        resymbol = []
        index = []
        for i,cell in enumerate(col_1):
            if cell.value == "no":
                index.append(i)
        for i,cell in enumerate(col_2):
            if i in index:
                resymbol.append(cell.value)
        print("-------共有{}个文件需要重新下载-------".format(len(resymbol)))
        print("# redownload # begin at:", time.ctime())
        for country in countries:
            repeat = 3
            print("\n---", country, "start ---")
            urls_path = "./util/Owner_URLs_{}.json".format(country)
            with open(urls_path, "r", encoding="utf-8") as f:
                urls = json.load(f)  # dict

            thread_list = []

            for symbol in resymbol:

                if urls[symbol] != '':
                    thread_list.append(GetMorningStarData(symbol=symbol, base_url=urls[symbol], country=country))

            print("Downloading", len(thread_list))

            for i, thread in enumerate(thread_list):
                thread.name = thread.symbol
                    # time.sleep(2)
                if i % 5 == 0 and i != 0:
                    time.sleep(10)
                if i % 10 == 0 and i != 0:
                    time.sleep(30)
                if i % 100 == 0 and i != 0:
                    time.sleep(60)
                print("=== {} start threading ===".format(thread.symbol))
                thread.start()
            for thread in thread_list:
                thread.join()
        print("=======下载完毕========")
        print("# Download # finish at:", time.ctime())