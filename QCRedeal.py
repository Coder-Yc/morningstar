import os
import time
import re
import threading
# from classes.QCDataClass_vs2 import QCData
import pandas as pd
import openpyxl

import re
import threading
import time
from threading import Thread
from pandas import DataFrame, read_excel, ExcelWriter, isna, Index
from openpyxl import load_workbook
from openpyxl.styles import Alignment


Quarters = {
    "1": 1, "2": 1, "3": 1,
    "4": 2, "5": 2, "6": 2,
    "7": 3, "8": 3, "9": 3,
    "10": 4, "11": 4, "12": 4
}

Frequency = "Quarterly"


class QCData(Thread):
    def __init__(self, country, excel_file, excel_path):
        super(QCData, self).__init__()
        # pre-settings
        self.country = country
        self.symbol = re.sub(r'(Download_)|(\.xlsx)', "", excel_file)  # company Name
        self.data_path = excel_path + excel_file
        self.added_heads = ["Symbol", "Date", "Frequency", "Name"]
        self.funds_heads = ["TotalSharesHeld", "TotalAssets", "CurrentShares", "ChangeAmount", "ChangePercentage",
                            "StarRating"]
        self.institutions_heads = ["TotalSharesHeld", "TotalAssets", "CurrentShares", "ChangeAmount"]
        self.labels = ["OwnershipData", "ConcentratedOwners", "Buyers", "Sellers"]
        self.labels_1 = ["OwnershipData", "ConcentratedOwners", "Buyers", "Sellers"]
        self.funds_heads_adj = []
        self.funds_heads_adj_1 = []
        self.institutions_heads_adj = []
        self.institutions_heads_adj_1 = []
        self.save_path = "./QCReport/{}/QCReport_{}.xlsx".format(self.country, self.symbol)

        # for datas
        self.funds_raw_data = {}
        self.institutions_raw_data = {}
        self.funds_QC_data = {}
        self.institutions_QC_data = {}
        # create head titles
        self.create_heads()
        self.null_symbol = False

    def create_heads(self):

        x1 = pd.ExcelFile(excel_path + excel_file)
        data = x1.sheet_names

        if 'Funds_OwnershipData' not in data:
            self.labels.remove('OwnershipData')
        if 'Funds_ConcentratedOwners' not in data:
            self.labels.remove('ConcentratedOwners')
        if 'Funds_Buyers' not in data:
            self.labels.remove('Buyers')
        if 'Funds_Sellers' not in data:
            self.labels.remove('Sellers')
        heads_adj = []
        for label in self.labels:
            for head in self.funds_heads:
                heads_adj.append(label + "_" + head)
        self.funds_heads_adj = self.added_heads + heads_adj

        if 'institutions_OwnershipData' not in data:
            self.labels_1.remove('OwnershipData')
        if 'institutions_ConcentratedOwners' not in data:
            self.labels_1.remove('ConcentratedOwners')
        if 'institutions_Buyers' not in data:
            self.labels_1.remove('Buyers')
        if 'institutions_Sellers' not in data:
            self.labels_1.remove('Sellers')
        # institutions
        heads_adj = []
        for label in self.labels_1:
            for head in self.institutions_heads:
                heads_adj.append(label + "_" + head)
        self.institutions_heads_adj = self.added_heads + heads_adj

    def switch_data(self):
        sheet_names = load_workbook(self.data_path).sheetnames
        if sheet_names == ["Sheet"]:
            self.null_symbol = True
            return
        for sheet_name in sheet_names:
            # funds
            if "Funds" in sheet_name:
                self.funds_raw_data[sheet_name] = read_excel(self.data_path, sheet_name=sheet_name)
                if self.funds_raw_data[sheet_name].empty:
                    col = self.funds_raw_data[sheet_name].columns.to_list()
                    col[0] = "Name"
                    self.funds_raw_data[sheet_name].columns = col
            # institutions
            elif "Institutions" in sheet_name:
                self.institutions_raw_data[sheet_name] = read_excel(self.data_path, sheet_name=sheet_name)
                if self.institutions_raw_data[sheet_name].empty:
                    col = self.institutions_raw_data[sheet_name].columns.to_list()
                    col[0] = "Name"
                    self.institutions_raw_data[sheet_name].columns = col

    def reformat_data(self):
        if self.null_symbol:
            return
        # Symbol Date Frequency Name %TotalSharesHeld %TotalAssets CurrentShares ChangeAmount Change% StarRating Trend
        for source_obj in self.funds_raw_data:
            dates_trends = self.funds_raw_data[source_obj].loc[:, ["Date", "Trend"]]
            current_shares = self.funds_raw_data[source_obj].loc[:, "CurrentShares"]
            df_temp = DataFrame(columns=self.funds_raw_data[source_obj].columns)
            for m in range(dates_trends.shape[0]):
                date = dates_trends.loc[m, :]["Date"]
                trends = dates_trends.loc[m, :]["Trend"]
                if isna(date):
                    continue
                year = date.split("-")[0]
                month = int(date.split("-")[1])
                if isna(trends):
                    temp = list()
                    temp.append("0")
                    trends = temp
                elif "," not in str(trends):
                    temp = list()
                    temp.append(trends)
                    trends = temp
                else:
                    trends = trends.split(",")
                trends.reverse()
                date_trend_adj = {}
                year_temp = int(year)
                quarter = Quarters[str(month)]
                for index, trend in enumerate(trends):
                    if index == 0:
                        date_trend_adj[year + "Q" + str(quarter)] = current_shares[m]
                    else:
                        if quarter == 0:
                            year_temp -= 1
                            quarter = 4
                            date_trend_adj[str(year_temp) + "Q4"] = trend
                        else:
                            date_trend_adj[str(year_temp) + "Q" + str(quarter)] = trend
                    quarter -= 1
                for index, item in enumerate(date_trend_adj.items()):
                    if index == 0:
                        s = self.funds_raw_data[source_obj].loc[m, :].copy()
                        s.loc["Date"] = item[0]
                        s.loc["CurrentShares"] = item[1]
                    else:
                        s = DataFrame(columns=self.funds_raw_data[source_obj].columns)
                        s.loc[0, "Name"] = self.funds_raw_data[source_obj].loc[m, "Name"]
                        s.loc[0, "Date"] = item[0]
                        s.loc[0, "CurrentShares"] = item[1]
                    df_temp = df_temp.append(s, ignore_index=True)
            self.funds_QC_data[source_obj] = df_temp

        # Symbol Date Frequency Name %TotalSharesHeld %TotalAssets CurrentShares ChangeAmount Trend
        for source_obj in self.institutions_raw_data:
            dates_trends = self.institutions_raw_data[source_obj].loc[:, ["Date", "Trend"]]

            # yangchong:添加urrent_shares
            current_shares = self.institutions_raw_data[source_obj].loc[:, "CurrentShares"]
            df_temp = DataFrame(columns=self.institutions_raw_data[source_obj].columns)
            for m in range(dates_trends.shape[0]):
                date = dates_trends.loc[m, :]["Date"]
                trends = dates_trends.loc[m, :]["Trend"]
                if isna(date):
                    continue
                date = dates_trends.loc[m, :]["Date"]
                year = date.split("-")[0]
                month = int(date.split("-")[1])
                if isna(trends):
                    temp = list()
                    temp.append("0")
                    trends = temp
                elif "," not in str(trends):
                    temp = list()
                    temp.append(trends)
                    trends = temp
                else:
                    # yangchong： 按照trends.split(" ")分开
                    trends = trends.split(" ")
                trends.reverse()
                date_trend_adj = {}
                year_temp = int(year)
                quarter = Quarters[str(month)]
                for index, trend in enumerate(trends):

                    # yangchong:修改为

                    if index == 0:
                        # yangchong：修改trend为 current_shares[m]
                        # date_trend_adj[year + "Q" + str(quarter)] = trend
                        date_trend_adj[year + "Q" + str(quarter)] = current_shares[m]
                    else:
                        if quarter == 0:
                            year_temp -= 1
                            quarter = 4
                            date_trend_adj[str(year_temp) + "Q4"] = trend
                        else:
                            date_trend_adj[str(year_temp) + "Q" + str(quarter)] = trend
                    quarter -= 1
                for index, item in enumerate(date_trend_adj.items()):
                    if index == 0:
                        s = self.institutions_raw_data[source_obj].loc[m, :].copy()
                        s.loc["Date"] = item[0]
                        s.loc["CurrentShares"] = item[1]
                    else:
                        s = DataFrame(columns=self.institutions_raw_data[source_obj].columns)
                        s.loc[0, "Name"] = self.institutions_raw_data[source_obj].loc[m, "Name"]
                        s.loc[0, "Date"] = item[0]
                        s.loc[0, "CurrentShares"] = item[1]
                    df_temp = df_temp.append(s, ignore_index=True)
            self.institutions_QC_data[source_obj] = df_temp

    def save_data(self):
        if self.null_symbol:
            return
        funds_save_df = DataFrame()
        institutions_save_df = DataFrame()
        for source_obj in self.funds_QC_data:
            obj = source_obj.split("_")[1]
            columns = {}
            for col in self.funds_QC_data[source_obj].columns:
                if col not in ["Name", "Date"]:
                    columns[col] = obj + "_" + col
            funds_save_df = funds_save_df.append(
                self.funds_QC_data[source_obj].rename(columns=columns))

        for source_obj in self.institutions_QC_data:
            obj = source_obj.split("_")[1]
            columns = {}
            for col in self.institutions_QC_data[source_obj].columns:
                if col not in ["Name", "Date"]:
                    columns[col] = obj + "_" + col
            institutions_save_df = institutions_save_df.append(
                self.institutions_QC_data[source_obj].rename(columns=columns))

        funds_save_df.insert(loc=0, column="Symbol", value=self.symbol)
        funds_save_df.insert(loc=1, column="Frequency", value=Frequency)
        institutions_save_df.insert(loc=0, column="Symbol", value=self.symbol)
        institutions_save_df.insert(loc=1, column="Frequency", value=Frequency)

        # funds_save_df_headers = [column for column in funds_save_df]
        # for col in funds_save_df_headers:
        #     if col in self.funds_heads_adj:
        #         self.funds_heads_adj_1.append(col)
        # self.funds_heads_adj_1.remove('Date')
        # if self.funds_heads_adj_1[1] != 'Date':
        #     self.funds_heads_adj_1.insert(1, 'Date')
        #
        # institutions_save_df_headers = [column for column in institutions_save_df]
        # for col in institutions_save_df_headers:
        #     if col in self.institutions_heads_adj:
        #         self.institutions_heads_adj_1.append(col)
        # self.institutions_heads_adj_1.remove('Date')
        # if self.institutions_heads_adj_1[1] != 'Date':
        #     self.institutions_heads_adj_1.insert(1, 'Date')

        with ExcelWriter(self.save_path, engine="openpyxl") as writer:
            funds_save_df.to_excel(writer, sheet_name="Funds", columns=self.funds_heads_adj, index=False)
            institutions_save_df.to_excel(writer, sheet_name="Institutions", columns=self.institutions_heads_adj, index=False)
            wb = writer.book
            for cell in wb["Funds"][1]:
                cell.alignment = Alignment(wrap_text=True)
            for cell in wb["Institutions"][1]:
                cell.alignment = Alignment(wrap_text=True)
            wb["Funds"].freeze_panes = "E2"
            wb["Institutions"].freeze_panes = "E2"
            writer.book = wb
            writer.save()


    def run(self):
        # yangchong：做一个pool_sema
        with pool_sema:
            self.switch_data()
            self.reformat_data()
            self.save_data()
            print("--- {} Done {} ---".format(self.symbol, time.ctime()))




if __name__ == "__main__":
    # 默认文件小于3k以及表格数量不正确的文件
    # flag_1代表文件大小，flag_2代表文件表数量
    flag_1=True
    flag_2=False

    countries = ['Shanghai_Shenzhen']
    FIRST_PATH = "QCReport/"
    print(time.ctime())
    if not os.path.exists(FIRST_PATH):
        os.mkdir(FIRST_PATH)
    # yangchong: 添加最大信号值
    max_connections = 8
    pool_sema = threading.BoundedSemaphore(max_connections)
    for country in countries:
        print("\n---", country, "start ---")
        SECOND_PATH = FIRST_PATH + "{}".format(country)
        if not os.path.exists(SECOND_PATH):
            os.mkdir(SECOND_PATH)
        thread_list = []
        # list files in directory 'RawData'
        excel_path = "./RawData/{}/".format(country)
        excel_files = os.listdir(excel_path)
        # 重新获取QC里的文件
        excel_path_QC = "./QCReport/{}/".format(country)
        excel_files_QC = os.listdir(excel_path_QC)
        excel_file_QC1 = []
        # 文件大小小于3kb的文件

        while flag_1:
                for excel_file_QC in excel_files_QC:
                    file_size = os.path.getsize("/Users/yangchong/Desktop/work/morningstar/QCReport/{}/{}".format(country,excel_file_QC))
                    if not file_size >= 3275:
                        excel_file_QC1.append(excel_file_QC)
                    flag_1 = False

            # 文件的表数量没有两个的文件
        while flag_2:
                for excel_file_QC in excel_files_QC:
                    file =  openpyxl.load_workbook("/Users/yangchong/Desktop/work/morningstar/QCReport/{}/{}".format(country,excel_file_QC))
                    file_sheetNum = len(file.sheetnames)
                    if file_sheetNum <=1 :
                        excel_file_QC1.append(excel_file_QC)
                    flag_2 = False
        print("共有{}个文件不符合".format(len(excel_file_QC1)))
    excel_file_QC1 = ['QCReport_900927.xlsx', 'QCReport_300306.xlsx', 'QCReport_002694.xlsx', 'QCReport_002552.xlsx', 'QCReport_300490.xlsx', 'QCReport_300644.xlsx', 'QCReport_002339.xlsx', 'QCReport_300351.xlsx', 'QCReport_300582.xlsx', 'QCReport_603088.xlsx', 'QCReport_300652.xlsx', 'QCReport_300717.xlsx', 'QCReport_603067.xlsx', 'QCReport_002006.xlsx', 'QCReport_002728.xlsx', 'QCReport_002682.xlsx', 'QCReport_002401.xlsx', 'QCReport_300469.xlsx', 'QCReport_300486.xlsx', 'QCReport_600470.xlsx', 'QCReport_600889.xlsx', 'QCReport_000980.xlsx', 'QCReport_600237.xlsx', 'QCReport_000502.xlsx', 'QCReport_603227.xlsx', 'QCReport_000017.xlsx', 'QCReport_300412.xlsx', 'QCReport_002207.xlsx', 'QCReport_600333.xlsx', 'QCReport_603266.xlsx', 'QCReport_603773.xlsx', 'QCReport_600671.xlsx', 'QCReport_300392.xlsx', 'QCReport_002529.xlsx', 'QCReport_002196.xlsx', 'QCReport_300189.xlsx', 'QCReport_300536.xlsx', 'QCReport_600939.xlsx', 'QCReport_002227.xlsx', 'QCReport_002620.xlsx', 'QCReport_600651.xlsx', 'QCReport_300062.xlsx', 'QCReport_000037.xlsx', 'QCReport_601113.xlsx', 'QCReport_002323.xlsx', 'QCReport_603315.xlsx', 'QCReport_002231.xlsx', 'QCReport_603895.xlsx', 'QCReport_603196.xlsx', 'QCReport_600684.xlsx', 'QCReport_002134.xlsx', 'QCReport_300449.xlsx', 'QCReport_300275.xlsx', 'QCReport_603790.xlsx', 'QCReport_600780.xlsx', 'QCReport_603006.xlsx', 'QCReport_000518.xlsx', 'QCReport_600807.xlsx', 'QCReport_300721.xlsx', 'QCReport_002319.xlsx', 'QCReport_300234.xlsx', 'QCReport_002748.xlsx', 'QCReport_300665.xlsx', 'QCReport_603115.xlsx', 'QCReport_603683.xlsx', 'QCReport_600555.xlsx', 'QCReport_600781.xlsx', 'QCReport_000722.xlsx', 'QCReport_300262.xlsx', 'QCReport_600152.xlsx', 'QCReport_603011.xlsx', 'QCReport_600778.xlsx', 'QCReport_000558.xlsx', 'QCReport_300018.xlsx', 'QCReport_000626.xlsx', 'QCReport_300809.xlsx', 'QCReport_002162.xlsx', 'QCReport_600113.xlsx', 'QCReport_600241.xlsx', 'QCReport_000962.xlsx', 'QCReport_600187.xlsx', 'QCReport_600353.xlsx', 'QCReport_002772.xlsx', 'QCReport_600095.xlsx', 'QCReport_300126.xlsx', 'QCReport_000523.xlsx', 'QCReport_603656.xlsx', 'QCReport_600345.xlsx', 'QCReport_002334.xlsx', 'QCReport_600083.xlsx', 'QCReport_600579.xlsx', 'QCReport_000020.xlsx', 'QCReport_600257.xlsx', 'QCReport_002363.xlsx', 'QCReport_300022.xlsx', 'QCReport_300472.xlsx', 'QCReport_300167.xlsx', 'QCReport_300537.xlsx', 'QCReport_600892.xlsx', 'QCReport_300405.xlsx', 'QCReport_603725.xlsx', 'QCReport_000779.xlsx', 'QCReport_600365.xlsx', 'QCReport_300239.xlsx', 'QCReport_300669.xlsx', 'QCReport_300517.xlsx', 'QCReport_000407.xlsx', 'QCReport_300281.xlsx', 'QCReport_002656.xlsx', 'QCReport_300444.xlsx', 'QCReport_600518.xlsx', 'QCReport_600148.xlsx', 'QCReport_000692.xlsx', 'QCReport_300278.xlsx', 'QCReport_600774.xlsx', 'QCReport_000153.xlsx', 'QCReport_300106.xlsx', 'QCReport_002494.xlsx', 'QCReport_603226.xlsx', 'QCReport_002752.xlsx', 'QCReport_600723.xlsx', 'QCReport_002247.xlsx', 'QCReport_000780.xlsx', 'QCReport_603318.xlsx', 'QCReport_600888.xlsx', 'QCReport_603031.xlsx', 'QCReport_600830.xlsx', 'QCReport_603089.xlsx', 'QCReport_300583.xlsx', 'QCReport_002504.xlsx', 'QCReport_300700.xlsx', 'QCReport_300350.xlsx', 'QCReport_002046.xlsx', 'QCReport_002553.xlsx', 'QCReport_300491.xlsx', 'QCReport_603861.xlsx', 'QCReport_000702.xlsx', 'QCReport_300025.xlsx', 'QCReport_000420.xlsx', 'QCReport_300475.xlsx', 'QCReport_300619.xlsx', 'QCReport_002671.xlsx', 'QCReport_300249.xlsx', 'QCReport_300588.xlsx', 'QCReport_300064.xlsx', 'QCReport_300434.xlsx', 'QCReport_600568.xlsx', 'QCReport_002149.xlsx', 'QCReport_002630.xlsx', 'QCReport_600354.xlsx', 'QCReport_000573.xlsx', 'QCReport_002667.xlsx', 'QCReport_002688.xlsx', 'QCReport_002722.xlsx', 'QCReport_002535.xlsx', 'QCReport_002165.xlsx', 'QCReport_600817.xlsx', 'QCReport_002470.xlsx', 'QCReport_002098.xlsx', 'QCReport_002077.xlsx', 'QCReport_002574.xlsx', 'QCReport_603329.xlsx', 'QCReport_600856.xlsx', 'QCReport_300265.xlsx', 'QCReport_300635.xlsx', 'QCReport_002173.xlsx', 'QCReport_000908.xlsx', 'QCReport_600381.xlsx', 'QCReport_000637.xlsx', 'QCReport_300615.xlsx', 'QCReport_600749.xlsx', 'QCReport_603165.xlsx', 'QCReport_000586.xlsx', 'QCReport_002554.xlsx', 'QCReport_600122.xlsx', 'QCReport_600821.xlsx', 'QCReport_600358.xlsx', 'QCReport_300654.xlsx', 'QCReport_300711.xlsx', 'QCReport_002779.xlsx', 'QCReport_300746.xlsx', 'QCReport_900937.xlsx', 'QCReport_300195.xlsx', 'QCReport_300480.xlsx', 'QCReport_600476.xlsx', 'QCReport_600661.xlsx', 'QCReport_300382.xlsx', 'QCReport_600724.xlsx', 'QCReport_300228.xlsx', 'QCReport_300414.xlsx', 'QCReport_000695.xlsx', 'QCReport_002094.xlsx', 'QCReport_603333.xlsx', 'QCReport_000416.xlsx', 'QCReport_300269.xlsx', 'QCReport_002082.xlsx', 'QCReport_000545.xlsx', 'QCReport_002578.xlsx', 'QCReport_600509.xlsx', 'QCReport_002256.xlsx', 'QCReport_002606.xlsx', 'QCReport_002190.xlsx', 'QCReport_300402.xlsx', 'QCReport_000007.xlsx', 'QCReport_300116.xlsx', 'QCReport_603689.xlsx', 'QCReport_600226.xlsx', 'QCReport_002742.xlsx', 'QCReport_300141.xlsx', 'QCReport_002715.xlsx', 'QCReport_002650.xlsx', 'QCReport_300268.xlsx', 'QCReport_300157.xlsx', 'QCReport_002187.xlsx', 'QCReport_603109.xlsx', 'QCReport_600375.xlsx', 'QCReport_002406.xlsx', 'QCReport_600861.xlsx', 'QCReport_300317.xlsx', 'QCReport_600836.xlsx', 'QCReport_002451.xlsx', 'QCReport_002514.xlsx', 'QCReport_300086.xlsx', 'QCReport_002144.xlsx', 'QCReport_002778.xlsx', 'QCReport_002447.xlsx', 'QCReport_000929.xlsx', 'QCReport_002502.xlsx', 'QCReport_300643.xlsx', 'QCReport_000616.xlsx', 'QCReport_300356.xlsx', 'QCReport_600877.xlsx', 'QCReport_300478.xlsx', 'QCReport_000587.xlsx', 'QCReport_002040.xlsx', 'QCReport_600898.xlsx', 'QCReport_603021.xlsx', 'QCReport_002693.xlsx', 'QCReport_300301.xlsx', 'QCReport_601968.xlsx', 'QCReport_900957.xlsx', 'QCReport_600695.xlsx', 'QCReport_002172.xlsx', 'QCReport_002037.xlsx', 'QCReport_300321.xlsx', 'QCReport_002719.xlsx', 'QCReport_600292.xlsx', 'QCReport_000548.xlsx', 'QCReport_002575.xlsx', 'QCReport_300337.xlsx', 'QCReport_900916.xlsx', 'QCReport_600791.xlsx', 'QCReport_000698.xlsx', 'QCReport_000677.xlsx', 'QCReport_002076.xlsx', 'QCReport_600396.xlsx', 'QCReport_300730.xlsx', 'QCReport_000509.xlsx', 'QCReport_300419.xlsx', 'QCReport_002471.xlsx', 'QCReport_600617.xlsx', 'QCReport_300032.xlsx', 'QCReport_000821.xlsx', 'QCReport_600355.xlsx', 'QCReport_002774.xlsx', 'QCReport_600139.xlsx', 'QCReport_002762.xlsx', 'QCReport_603703.xlsx', 'QCReport_300589.xlsx', 'QCReport_000533.xlsx', 'QCReport_600969.xlsx', 'QCReport_603168.xlsx', 'QCReport_600744.xlsx', 'QCReport_002735.xlsx', 'QCReport_600251.xlsx', 'QCReport_000421.xlsx', 'QCReport_000837.xlsx', 'QCReport_002109.xlsx', 'QCReport_002559.xlsx', 'QCReport_000635.xlsx', 'QCReport_603110.xlsx', 'QCReport_002464.xlsx', 'QCReport_300549.xlsx', 'QCReport_000727.xlsx', 'QCReport_002433.xlsx', 'QCReport_600157.xlsx', 'QCReport_002576.xlsx', 'QCReport_300621.xlsx', 'QCReport_002219.xlsx', 'QCReport_900915.xlsx', 'QCReport_300733.xlsx', 'QCReport_600815.xlsx', 'QCReport_000530.xlsx', 'QCReport_300135.xlsx', 'QCReport_600593.xlsx', 'QCReport_002691.xlsx', 'QCReport_900922.xlsx', 'QCReport_300753.xlsx', 'QCReport_000993.xlsx', 'QCReport_300211.xlsx', 'QCReport_600822.xlsx', 'QCReport_002280.xlsx', 'QCReport_300084.xlsx', 'QCReport_300591.xlsx', 'QCReport_600137.xlsx', 'QCReport_600834.xlsx', 'QCReport_601008.xlsx', 'QCReport_600619.xlsx', 'QCReport_600249.xlsx', 'QCReport_300600.xlsx', 'QCReport_000985.xlsx', 'QCReport_002054.xlsx', 'QCReport_000593.xlsx', 'QCReport_600530.xlsx', 'QCReport_002613.xlsx', 'QCReport_002490.xlsx', 'QCReport_300786.xlsx', 'QCReport_900918.xlsx', 'QCReport_300293.xlsx', 'QCReport_000045.xlsx', 'QCReport_603330.xlsx', 'QCReport_000803.xlsx', 'QCReport_300505.xlsx', 'QCReport_300155.xlsx', 'QCReport_002652.xlsx', 'QCReport_002255.xlsx', 'QCReport_000792.xlsx', 'QCReport_002469.xlsx', 'QCReport_300051.xlsx', 'QCReport_300817.xlsx', 'QCReport_300114.xlsx', 'QCReport_300499.xlsx', 'QCReport_300163.xlsx', 'QCReport_000835.xlsx', 'QCReport_600896.xlsx', 'QCReport_000862.xlsx', 'QCReport_000927.xlsx', 'QCReport_002275.xlsx', 'QCReport_300067.xlsx', 'QCReport_603828.xlsx', 'QCReport_002633.xlsx', 'QCReport_002263.xlsx', 'QCReport_000570.xlsx', 'QCReport_300175.xlsx', 'QCReport_600615.xlsx', 'QCReport_000659.xlsx', 'QCReport_000788.xlsx', 'QCReport_000622.xlsx', 'QCReport_300698.xlsx', 'QCReport_900943.xlsx', 'QCReport_000767.xlsx', 'QCReport_603042.xlsx', 'QCReport_000419.xlsx', 'QCReport_002561.xlsx', 'QCReport_300270.xlsx', 'QCReport_600506.xlsx', 'QCReport_002062.xlsx', 'QCReport_000663.xlsx', 'QCReport_600290.xlsx', 'QCReport_600785.xlsx', 'QCReport_002520.xlsx', 'QCReport_900955.xlsx']
    for excel_file in excel_file_QC1:
        excel_file = excel_file.replace('QCReport','Download')
        thread_list.append(QCData(country, excel_file, excel_path))
    for num, thread in enumerate(thread_list):
        thread.name = str(num) + " " + thread.symbol
        print("=== {} start threading ===".format(thread.symbol))
        thread.start()
    for thread in thread_list:
        thread.join()
    print(time.ctime())



