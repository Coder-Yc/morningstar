import os
import re
import threading
import time
from threading import Thread
import pandas as pd
from pandas import DataFrame, read_excel, ExcelWriter, isna, Index
from openpyxl import load_workbook
from openpyxl.styles import Alignment


Quarters = {
    "1": 1, "2": 1, "3": 1,
    "4": 2, "5": 2, "6": 2,
    "7": 3, "8": 3, "9": 3,
    "10": 4, "11": 4, "12": 4
}

Frequency = "Quarterly"

# 定义一个清洗类对download下来的数据进行清洗
# 把下载下来的表格分为funds和Institutions以及把trend分给到每个季度（清洗的内容具体参考跑出来的结果）
# 值得注意的是这里的线程不好控制，容易线程开太多爆掉，建议每个国家分批次允许
#   Output: ./QCReport/"Country"/QCReport_"xx".xlsx
class QCData(Thread):
    def __init__(self, country, excel_file, excel_path):
        super(QCData, self).__init__()
        # pre-settings
        self.country = country
        self.symbol = re.sub(r'(Download_)|(\.xlsx)', "", excel_file)  # company Name
        self.data_path = excel_path + excel_file
        self.added_heads = ["Symbol", "Date", "Frequency", "Name"]
        self.funds_heads = ["TotalSharesHeld", "TotalAssets", "CurrentShares", "ChangeAmount", "ChangePercentage",
                            "StarRating"]
        self.institutions_heads = ["TotalSharesHeld", "TotalAssets", "CurrentShares", "ChangeAmount"]
        self.labels = ["OwnershipData", "ConcentratedOwners", "Buyers", "Sellers"]
        self.labels_1 = ["OwnershipData", "ConcentratedOwners", "Buyers", "Sellers"]
        self.funds_heads_adj = []
        self.funds_heads_adj_1 = []
        self.institutions_heads_adj = []
        self.institutions_heads_adj_1 = []
        self.save_path = "./QCReport/{}/QCReport_{}.xlsx".format(self.country, self.symbol)

        # for datas
        self.funds_raw_data = {}
        self.institutions_raw_data = {}
        self.funds_QC_data = {}
        self.institutions_QC_data = {}
        # create head titles
        self.create_heads()
        self.null_symbol = False

    def create_heads(self):
        # funds
        x1 = pd.ExcelFile(excel_path+excel_file)
        data = x1.sheet_names

        if 'Funds_OwnershipData' not in data:
            self.labels.remove('OwnershipData')
        if 'Funds_ConcentratedOwners' not in data:
            self.labels.remove('ConcentratedOwners')
        if 'Funds_Buyers' not in data:
            self.labels.remove('Buyers')
        if 'Funds_Sellers' not in data:
            self.labels.remove('Sellers')
        heads_adj = []
        for label in self.labels:
            for head in self.funds_heads:
                heads_adj.append(label + "_" + head)
        self.funds_heads_adj = self.added_heads + heads_adj

        if 'Institutions_OwnershipData' not in data:
            self.labels_1.remove('OwnershipData')
        if 'Institutions_ConcentratedOwners' not in data:
            self.labels_1.remove('ConcentratedOwners')
        if 'Institutions_Buyers' not in data:
            self.labels_1.remove('Buyers')
        if 'Institutions_Sellers' not in data:
            self.labels_1.remove('Sellers')
        # institutions
        heads_adj = []
        for label in self.labels_1:
            for head in self.institutions_heads:
                heads_adj.append(label + "_" + head)
        self.institutions_heads_adj = self.added_heads + heads_adj

    def switch_data(self):
        sheet_names = load_workbook(self.data_path).sheetnames
        if sheet_names == ["Sheet"]:
            self.null_symbol = True
            return
        for sheet_name in sheet_names:
            # funds
            if "Funds" in sheet_name:
                self.funds_raw_data[sheet_name] = read_excel(self.data_path, sheet_name=sheet_name)
                if self.funds_raw_data[sheet_name].empty:
                    col = self.funds_raw_data[sheet_name].columns.to_list()
                    col[0] = "Name"
                    self.funds_raw_data[sheet_name].columns = col
            # institutions
            elif "Institutions" in sheet_name:
                self.institutions_raw_data[sheet_name] = read_excel(self.data_path, sheet_name=sheet_name)
                if self.institutions_raw_data[sheet_name].empty:
                    col = self.institutions_raw_data[sheet_name].columns.to_list()
                    col[0] = "Name"
                    self.institutions_raw_data[sheet_name].columns = col

    def reformat_data(self):
        if self.null_symbol:
            return
        # Symbol Date Frequency Name %TotalSharesHeld %TotalAssets CurrentShares ChangeAmount Change% StarRating Trend
        for source_obj in self.funds_raw_data:
            dates_trends = self.funds_raw_data[source_obj].loc[:, ["Date", "Trend"]]
            current_shares = self.funds_raw_data[source_obj].loc[:, "CurrentShares"]
            df_temp = DataFrame(columns=self.funds_raw_data[source_obj].columns)
            for m in range(dates_trends.shape[0]):
                date = dates_trends.loc[m, :]["Date"]
                trends = dates_trends.loc[m, :]["Trend"]
                if isna(date):
                    continue
                year = date.split("-")[0]
                month = int(date.split("-")[1])
                if isna(trends):
                    temp = list()
                    temp.append("0")
                    trends = temp
                elif " " not in str(trends):
                    temp = list()
                    temp.append(trends)
                    trends = temp
                else:
                    trends = trends.split(" ")
                trends.reverse()
                date_trend_adj = {}
                year_temp = int(year)
                quarter_current = Quarters[str(month)]
                for index, trend in enumerate(trends):
                    if index == 0:
                        date_trend_adj[year + "Q" + str(quarter_current)] = current_shares[m]
                    else:
                        if quarter_current == 0:
                            year_temp -= 1

                            quarter_current = 4
                            date_trend_adj[str(year_temp) + "Q4"] = trend
                        else:
                            date_trend_adj[str(year_temp) + "Q" + str(quarter_current)] = trend
                    quarter_current -= 1
                for index, item in enumerate(date_trend_adj.items()):
                    if index == 0:
                        s = self.funds_raw_data[source_obj].loc[m, :].copy()
                        s.loc["Date"] = item[0]
                        s.loc["CurrentShares"] = item[1]
                    else:
                        s = DataFrame(columns=self.funds_raw_data[source_obj].columns)
                        s.loc[0, "Name"] = self.funds_raw_data[source_obj].loc[m, "Name"]
                        s.loc[0, "Date"] = item[0]
                        s.loc[0, "CurrentShares"] = item[1]
                    df_temp = df_temp.append(s, ignore_index=True)
            self.funds_QC_data[source_obj] = df_temp

        # Symbol Date Frequency Name %TotalSharesHeld %TotalAssets CurrentShares ChangeAmount Trend
        for source_obj in self.institutions_raw_data:
            dates_trends = self.institutions_raw_data[source_obj].loc[:, ["Date", "Trend"]]
            current_shares = self.institutions_raw_data[source_obj].loc[:, "CurrentShares"]
            df_temp = DataFrame(columns=self.institutions_raw_data[source_obj].columns)
            for m in range(dates_trends.shape[0]):
                date = dates_trends.loc[m, :]["Date"]
                trends = dates_trends.loc[m, :]["Trend"]
                if isna(date):
                    continue
                date = dates_trends.loc[m, :]["Date"]
                year = date.split("-")[0]
                month = int(date.split("-")[1])
                if isna(trends):
                    temp = list()
                    temp.append("0")
                    trends = temp
                elif " " not in str(trends):
                    temp = list()
                    temp.append(trends)
                    trends = temp
                else:
                    trends = trends.split(" ")
                trends.reverse()
                date_trend_adj = {}
                year_temp = int(year)
                quarter = Quarters[str(month)]
                for index, trend in enumerate(trends):
                    if index == 0:
                        date_trend_adj[year + "Q" + str(quarter)] = current_shares[m]
                    else:
                        if quarter == 0:
                            year_temp -= 1
                            quarter = 4
                            date_trend_adj[str(year_temp) + "Q4"] = trend
                        else:
                            date_trend_adj[str(year_temp) + "Q" + str(quarter)] = trend
                    quarter -= 1
                for index, item in enumerate(date_trend_adj.items()):
                    if index == 0:
                        s = self.institutions_raw_data[source_obj].loc[m, :].copy()
                        s.loc["Date"] = item[0]
                        s.loc["CurrentShares"] = item[1]
                    else:
                        s = DataFrame(columns=self.institutions_raw_data[source_obj].columns)
                        s.loc[0, "Name"] = self.institutions_raw_data[source_obj].loc[m, "Name"]
                        s.loc[0, "Date"] = item[0]
                        s.loc[0, "CurrentShares"] = item[1]
                    df_temp = df_temp.append(s, ignore_index=True)
            self.institutions_QC_data[source_obj] = df_temp

    def save_data(self):
        if self.null_symbol:
            return
        funds_save_df = DataFrame()
        institutions_save_df = DataFrame()
        for source_obj in self.funds_QC_data:
            obj = source_obj.split("_")[1]
            columns = {}
            for col in self.funds_QC_data[source_obj].columns:
                if col not in ["Name", "Date"]:
                    columns[col] = obj + "_" + col
            funds_save_df = funds_save_df.append(self.funds_QC_data[source_obj].rename(columns=columns))

        for source_obj in self.institutions_QC_data:
            obj = source_obj.split("_")[1]
            columns = {}
            for col in self.institutions_QC_data[source_obj].columns:
                if col not in ["Name", "Date"]:
                    columns[col] = obj + "_" + col
            institutions_save_df = institutions_save_df.append(
            self.institutions_QC_data[source_obj].rename(columns=columns))

        funds_save_df.insert(loc=0, column="Symbol", value=self.symbol)
        funds_save_df.insert(loc=1, column="Frequency", value=Frequency)
        institutions_save_df.insert(loc=0, column="Symbol", value=self.symbol)
        institutions_save_df.insert(loc=1, column="Frequency", value=Frequency)

        with ExcelWriter(self.save_path, engine="openpyxl") as writer:
            funds_save_df.to_excel(writer, sheet_name="Funds", columns=self.funds_heads_adj, index=False)
            institutions_save_df.to_excel(writer, sheet_name="Institutions", columns=self.institutions_heads_adj, index=False)
            wb = writer.book
            for cell in wb["Funds"][1]:
                cell.alignment = Alignment(wrap_text=True)
            for cell in wb["Institutions"][1]:
                cell.alignment = Alignment(wrap_text=True)
            wb["Funds"].freeze_panes = "E2"
            wb["Institutions"].freeze_panes = "E2"
            writer.book = wb
            writer.save()
            # writer.close()

    def run(self):
        with pool_sema:
            self.switch_data()
            self.reformat_data()
            self.save_data()
            print("--- {} Done {} ---".format(self.symbol, time.ctime()))


if __name__ == '__main__':
    # 'TSX', 'Shanghai_Shenzhen','Snp500_Ru1000',
    countries = ['Snp500_Ru1000']
    FIRST_PATH = "QCReport/"
    print(time.ctime())
    if not os.path.exists(FIRST_PATH):
        os.mkdir(FIRST_PATH)
    max_connections = 8
    pool_sema = threading.BoundedSemaphore(max_connections)
    for country in countries:
        print("\n---", country, "start ---")
        SECOND_PATH = FIRST_PATH + "{}".format(country)
        if not os.path.exists(SECOND_PATH):
            os.mkdir(SECOND_PATH)
        thread_list = []
        # list files in directory 'RawData'
        excel_path = "./RawData/{}/".format(country)
        excel_files = os.listdir(excel_path)

        for excel_file in excel_files[:20]:
            print(excel_file)
            thread_list.append(QCData(country, excel_file, excel_path))
        for num, thread in enumerate(thread_list):
            thread.name = str(num) + " " + thread.symbol
            print("=== {} start threading ===".format(thread.symbol))
            thread.start()
        for thread in thread_list:
            thread.join()
    print(time.ctime())
